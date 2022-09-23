#encoding: utf-8
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as num2col
from openpyxl.utils import column_index_from_string as col2num
import sys 
reload(sys) 
sys.setdefaultencoding('utf-8')

wb1 = load_workbook(filename="yun.xlsx") 
ws1 = wb1.get_sheet_names()
s1 = wb1.get_sheet_by_name(ws1[0])

wb2 = load_workbook(filename="lims.xlsx") 
ws2 = wb2.get_sheet_names()
s2 = wb2.get_sheet_by_name(ws2[0])

wb3 = load_workbook(filename="module.xlsx")
ws3 = wb3.get_sheet_names()
s3 = wb3.get_sheet_by_name(ws3[0])
mrow = s2.max_row
mcol = s2.max_column
for m in range(2,mrow+1):
    for n in range(3,mcol): 
        #ascii a 97, till {|}~ not charactor
        n2=n-2
        n=num2col(n)
        i='%s%d'%(n,m)
        n2=num2col(n2)
        i2='%s%d'%(n2,m)
        tmp=s2[i].value
        s3[i2].value=tmp

#fun2
for m in range(2,mrow+1):
    n="AC"
    i='%s%d'%(n,m)
    tmp=s2[i].value
    if tmp == "未知3":
        i2='%s%d'%("AL",m)
        tmp2=s2[i2].value
        i3='%s%d'%("AI",m)
        s3[i3].value=tmp2
        i4='%s%d'%("AJ",m)
        s3[i4].value=""

#fun3 short
mrow3 = s3.max_row
mrow1 = s1.max_row
for m1 in range(4,mrow1+1):
    n = 0
    i1 = '%s%d'%("B",m1)
    u1 = s1[i1].value
    ig1 = '%s%d'%("K",m1)
    g1 = s1[ig1].value
    ir1 = '%s%d'%("M",m1)
    r1 = s1[ir1].value
    for m3 in range(2,mrow3+1):
        i3 = '%s%d'%("AA",m3)
        u3 = s3[i3].value
        if u1 == u3:
            ig3 = '%s%d'%("AJ",m3)
            s3[ig3].value = g1
            ir3 = '%s%d'%("AK",m3)
            s3[ir3].value = r1
            break 

wb3.save('hh.xlsx') 
wb3.close()